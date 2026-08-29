from fastapi import APIRouter, HTTPException, status, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, date
from ..database import supabase
from ..middleware.auth import require_faculty, CurrentUser
import io
import uuid
from openpyxl import Workbook, load_workbook
from ..routes.auth import get_password_hash
from .student import get_next_leave_status


router = APIRouter(prefix="/api/faculty", tags=["Faculty"])


# ==================== PYDANTIC MODELS ====================

class AttendanceEntry(BaseModel):
    """Single attendance entry."""
    student_id: str
    status: str  # present, absent, late


class AttendanceCreate(BaseModel):
    """Create attendance for a class."""
    class_year: str
    section: str
    subject: str
    date: date
    entries: List[AttendanceEntry]


class MarksEntry(BaseModel):
    """Single marks entry."""
    student_id: str
    marks_obtained: float


class MarksCreate(BaseModel):
    """Create marks for students."""
    class_year: str
    section: str
    subject: str
    exam_type: str  # internal1, internal2, internal3, external
    max_marks: float
    entries: List[MarksEntry]


class ProfileUpdate(BaseModel):
    """Update faculty profile."""
    mobile: Optional[str] = None
    availability_status: Optional[bool] = None


class StudentCreate(BaseModel):
    """Create a single student."""
    name: str
    register_number: str
    roll_number: str
    class_year: str
    section: str
    batch: str
    mobile: Optional[str] = None
    email: Optional[str] = None
    father_name: Optional[str] = None
    mother_name: Optional[str] = None



class StudentUpdate(BaseModel):
    """Update a student."""
    name: Optional[str] = None
    register_number: Optional[str] = None
    roll_number: Optional[str] = None
    mobile: Optional[str] = None
    father_name: Optional[str] = None
    mother_name: Optional[str] = None


class LeaveRequestCreate(BaseModel):
    """Create a leave request."""
    leave_type: str
    from_date: date
    to_date: date
    reason: str


class LeaveStatusUpdate(BaseModel):
    """Update leave status."""
    status: str  # approved, rejected
    rejection_reason: Optional[str] = None



# ==================== PROFILE ROUTES ====================

@router.get("/profile")
async def get_profile(current_user: CurrentUser = Depends(require_faculty)):
    """Get faculty profile."""
    try:
        result = supabase.table("faculty").select("*, users(email)").eq("id", current_user.id).execute()
        
        if not result.data:
            raise HTTPException(status_code=404, detail="Profile not found")
        
        f = result.data[0]
        return {
            "id": f["id"],
            "name": f["name"],
            "email": f["users"]["email"] if f.get("users") else "",
            "employee_id": f["employee_id"],
            "mobile": f.get("mobile"),
            "department": f.get("department"),
            "availability_status": f.get("availability_status", True)
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/profile")
async def update_profile(
    profile: ProfileUpdate,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Update faculty profile."""
    try:
        update_data = {k: v for k, v in profile.dict().items() if v is not None}
        result = supabase.table("faculty").update(update_data).eq("id", current_user.id).execute()
        return {"message": "Profile updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== CLASSES ROUTES ====================

@router.get("/classes")
async def get_assigned_classes(current_user: CurrentUser = Depends(require_faculty)):
    """Get classes assigned to this faculty."""
    try:
        result = supabase.table("faculty_classes").select("*, classes(*)").eq("faculty_id", current_user.id).execute()
        return {"classes": result.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/students")
async def get_students_by_class(
    class_year: str,
    section: str,
    batch: str = None,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Get students for a specific class."""
    try:
        query = supabase.table("students").select(
            "id, name, register_number, roll_number, batch"
        ).eq("class_year", class_year).eq("section", section)
        
        if batch:
            query = query.eq("batch", batch)
            
        result = query.order("roll_number").execute()
        
        return {"students": result.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== ATTENDANCE ROUTES ====================

@router.post("/attendance")
async def mark_attendance(
    attendance: AttendanceCreate,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Mark attendance for a class."""
    try:
        records = []
        for entry in attendance.entries:
            records.append({
                "student_id": entry.student_id,
                "faculty_id": current_user.id,
                "class_year": attendance.class_year,
                "section": attendance.section,
                "subject": attendance.subject,
                "date": attendance.date.isoformat(),
                "status": entry.status,
                "period": 1,
                "class_id": None,
                "subject_id": None
            })
        
        result = supabase.table("attendance").insert(records).execute()
        
        return {
            "message": "Attendance marked successfully",
            "count": len(records)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/attendance/{class_year}/{section}")
async def get_class_attendance(
    class_year: str,
    section: str,
    subject: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Get attendance records for a class."""
    try:
        query = supabase.table("attendance").select(
            "*, students(name, roll_number)"
        ).eq("class_year", class_year).eq("section", section)
        
        if subject:
            query = query.eq("subject", subject)
        if date_from:
            query = query.gte("date", date_from.isoformat())
        if date_to:
            query = query.lte("date", date_to.isoformat())
        
        result = query.order("date", desc=True).execute()
        
        return {"attendance": result.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/attendance/summary/{class_year}/{section}")
async def get_attendance_summary(
    class_year: str,
    section: str,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Get attendance summary with percentages for a class."""
    try:
        # Get all students in class
        students = supabase.table("students").select(
            "id, name, roll_number"
        ).eq("class_year", class_year).eq("section", section).execute()
        
        summary = []
        for student in students.data:
            attendance = supabase.table("attendance").select("status").eq("student_id", student["id"]).execute()
            total = len(attendance.data)
            present = sum(1 for a in attendance.data if a["status"] == "present")
            percentage = (present / total * 100) if total > 0 else 0
            
            summary.append({
                "student_id": student["id"],
                "name": student["name"],
                "roll_number": student["roll_number"],
                "total_classes": total,
                "present": present,
                "percentage": round(percentage, 2)
            })
        
        return {"summary": summary}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== MARKS ROUTES ====================

@router.post("/marks")
async def enter_marks(
    marks: MarksCreate,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Enter marks for students."""
    try:
        subject_result = supabase.table("subjects").select("id").ilike("name", marks.subject.strip()).limit(1).execute()
        subject_id = subject_result.data[0]["id"] if subject_result.data else None
        records = []
        for entry in marks.entries:
            records.append({
                "student_id": entry.student_id,
                "subject_id": subject_id,
                "subject": marks.subject.strip(),
                "exam_type": marks.exam_type,
                "max_marks": marks.max_marks,
                "marks_obtained": entry.marks_obtained
            })
        
        result = supabase.table("marks").insert(records).execute()
        
        return {
            "message": "Marks entered successfully",
            "count": len(records)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/marks/{marks_id}")
async def update_marks(
    marks_id: str,
    marks_obtained: float,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Update marks for a student."""
    try:
        result = supabase.table("marks").update({
            "marks_obtained": marks_obtained
        }).eq("id", marks_id).execute()
        
        if not result.data:
            raise HTTPException(status_code=404, detail="Marks record not found")
        
        return {"message": "Marks updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/marks/{class_year}/{section}")
async def get_class_marks(
    class_year: str,
    section: str,
    subject: Optional[str] = None,
    exam_type: Optional[str] = None,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Get marks for a class."""
    try:
        students_result = supabase.table("students").select("id").eq("class_year", class_year).eq("section", section).execute()
        student_ids = [student["id"] for student in (students_result.data or [])]

        if not student_ids:
            return {"marks": []}

        query = supabase.table("marks").select(
            "*, students(name, roll_number), subjects(name)"
        ).in_("student_id", student_ids)

        if exam_type:
            query = query.eq("exam_type", exam_type)

        result = query.execute()
        marks_data = result.data or []

        if subject:
            marks_data = [
                item for item in marks_data
                    if ((item.get("subjects") or {}).get("name") or item.get("subject")) == subject
            ]

        return {"marks": marks_data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== EVENTS ROUTES ====================

@router.get("/events")
async def get_events(current_user: CurrentUser = Depends(require_faculty)):
    """Get all events for faculty to view."""
    try:
        result = supabase.table("events").select("*").order("event_date", desc=True).execute()
        return {"events": result.data}
    except Exception as e:
        # Return empty events if table doesn't exist
        return {"events": []}


# ==================== STUDENT MANAGEMENT ROUTES ====================

@router.get("/students/template")
async def download_student_template(current_user: CurrentUser = Depends(require_faculty)):
    """Download Excel template for student upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Header row - matching user's expected format
    headers = ["S.No", "Roll No", "Reg No", "Name", "Mobile", "Email", "Father Name", "Mother Name"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Sample data row
    sample = [1, "01", "REG2024001", "John Doe", "9876543210", "john@example.com", "James Doe", "Mary Doe"]
    for col, value in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=value)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_template.xlsx"}
    )


@router.post("/students/upload")
async def upload_students(
    file: UploadFile = File(...),
    class_year: str = Form(...),
    section: str = Form(...),
    batch: str = Form(...),
    current_user: CurrentUser = Depends(require_faculty)
):
    """Upload students from Excel file."""
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")
        
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Parse rows (skip header) - Column order: S.No, Roll No, Reg No, Name, Mobile, Email, Father Name, Mother Name
        students = []
        users = []
        errors = []
        seen_emails = set()

        role_res = supabase.table("roles").select("id").eq("name", "student").execute()
        if not role_res.data:
            raise HTTPException(status_code=500, detail="Student role not configured in database")
        student_role_id = role_res.data[0]["id"]

        existing_users = supabase.table("users").select("email, username").execute()
        existing_user_emails = {record.get("email", "").lower() for record in existing_users.data if record.get("email")}
        existing_usernames = {record.get("username", "").lower() for record in existing_users.data if record.get("username")}
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            try:
                # Column mapping: 0=S.No (skip), 1=Roll No, 2=Reg No, 3=Name
                roll_number = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                register_number = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                name = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                mobile = str(row[4]).strip() if len(row) > 4 and row[4] is not None and str(row[4]).strip() else None
                email = str(row[5]).strip() if len(row) > 5 and row[5] is not None and str(row[5]).strip() else None
                father_name = str(row[6]).strip() if len(row) > 6 and row[6] is not None and str(row[6]).strip() else None
                mother_name = str(row[7]).strip() if len(row) > 7 and row[7] is not None and str(row[7]).strip() else None
                
                # Validate required fields
                if not name or not roll_number or not register_number:
                    errors.append(f"Row {idx}: Missing required fields (Name, Roll No, or Reg No)")
                    continue

                # Generate missing email if needed
                if not email:
                    email = f"{register_number.lower()}@student.college.edu"

                normalized_email = email.strip().lower()
                if normalized_email in seen_emails or normalized_email in existing_user_emails:
                    errors.append(f"Row {idx}: Duplicate email '{email}' found in file or database")
                    continue
                seen_emails.add(normalized_email)

                # Generate User ID and Password
                user_id = str(uuid.uuid4())
                base_username = f"std_{register_number.lower().replace(' ', '')}"
                username = base_username
                username_counter = 1
                while username.lower() in existing_usernames or username.lower() in {u.get('username', '').lower() for u in users}:
                    username = f"{base_username}_{username_counter}"
                    username_counter += 1
                existing_usernames.add(username.lower())
                # Default password is lowercased roll number for now (or a standard default)
                default_password = roll_number.lower() 
                password_hash = get_password_hash(default_password)
                
                user_entry = {
                    "id": user_id,
                    "email": normalized_email,
                    "username": username,
                    "password_hash": password_hash,
                    "role_id": student_role_id,
                    "created_at": datetime.now().isoformat()
                }

                student_entry = {
                    "id": user_id, # Link to user
                    "roll_number": roll_number,
                    "register_number": register_number,
                    "name": name,
                    "mobile": mobile,
                    "father_name": father_name,
                    "mother_name": mother_name,
                    "class_year": class_year,
                    "section": section,
                    "batch": batch,
                    "approval_status": "pending"
                }
                
                students.append(student_entry)
                users.append(user_entry)

            except Exception as row_error:
                errors.append(f"Row {idx}: {str(row_error)}")
        
        if not students:
            error_msg = "No valid student data found in Excel file."
            if errors:
                error_msg += " Errors: " + "; ".join(errors[:5])
            raise HTTPException(status_code=400, detail=error_msg)
        
        # Insert into database
        print(f"Attempting to insert {len(users)} users and {len(students)} students...")
        
        # 1. Insert Users first
        user_result = supabase.table("users").insert(users).execute()
        if not user_result.data:
             print("Warning: User insert returned no data, potentially succeeded if simplified return.")
        
        # 2. Insert Students
        result = supabase.table("students").insert(students).execute()
        print(f"Database insert result: {result}")
        
        # Check if insert was successful
        if not result.data:
            raise HTTPException(status_code=500, detail="Database insert failed - no data returned")
        
        response = {
            "message": f"Successfully uploaded {len(students)} students",
            "count": len(students),
            "batch": batch
        }
        
        if errors:
            response["warnings"] = errors[:10]  # Include first 10 errors as warnings
        
        return response
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_detail = f"Upload failed: {str(e)}. Traceback: {traceback.format_exc()}"
        print(error_detail)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/students/batches")
async def get_batches(
    class_year: Optional[str] = None,
    section: Optional[str] = None,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Get all unique batches with student counts."""
    try:
        query = supabase.table("students").select("batch, class_year, section")
        
        if class_year:
            query = query.eq("class_year", class_year)
        if section:
            query = query.eq("section", section)
        
        result = query.execute()
        
        # Group by batch
        batch_counts = {}
        for student in result.data:
            batch = student.get("batch") or "Unassigned"
            key = f"{student['class_year']}|{student['section']}|{batch}"
            if key not in batch_counts:
                batch_counts[key] = {
                    "batch": batch,
                    "class_year": student["class_year"],
                    "section": student["section"],
                    "count": 0
                }
            batch_counts[key]["count"] += 1
        
        return {"batches": list(batch_counts.values())}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/students/by-batch")
async def get_students_by_batch(
    class_year: str,
    section: str,
    batch: Optional[str] = None,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Get students filtered by batch."""
    try:
        query = supabase.table("students").select("*").eq("class_year", class_year).eq("section", section)
        
        if batch and batch != "null":
            query = query.eq("batch", batch)
        
        result = query.order("roll_number").execute()
        return {"students": result.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/students")
async def create_student(
    student: StudentCreate,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Create a single student."""
    try:
        data = student.dict()
        result = supabase.table("students").insert(data).execute()
        return {"message": "Student created successfully", "student": result.data[0] if result.data else None}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/students/{student_id}")
async def update_student(
    student_id: str,
    student: StudentUpdate,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Update a student."""
    try:
        update_data = {k: v for k, v in student.dict().items() if v is not None}
        if not update_data:
            raise HTTPException(status_code=400, detail="No fields to update")
        
        result = supabase.table("students").update(update_data).eq("id", student_id).execute()
        
        if not result.data:
            raise HTTPException(status_code=404, detail="Student not found")
        
        return {"message": "Student updated successfully", "student": result.data[0]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/students/{student_id}")
async def delete_student(
    student_id: str,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Delete a student."""
    try:
        result = supabase.table("students").delete().eq("id", student_id).execute()
        return {"message": "Student deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== LEAVE MANAGEMENT ROUTES ====================

@router.get("/student-leaves")
async def get_student_leaves(current_user: CurrentUser = Depends(require_faculty)):
    """Get student leave requests awaiting faculty approval."""
    try:
        result = supabase.table("leave_requests").select("*").eq("role", "student").eq("status", "pending_faculty").order("created_at", desc=True).execute()
        requests = result.data or []

        for item in requests:
            user_id = item.get("user_id")
            item["users"] = []
            item["students"] = []

            if user_id:
                user_result = supabase.table("users").select("email").eq("id", user_id).limit(1).execute()
                if user_result.data:
                    item["users"] = user_result.data

                student_result = supabase.table("students").select("name, roll_number, class_year, section").eq("id", user_id).execute()
                if student_result.data:
                    item["students"] = student_result.data
                    item["name"] = student_result.data[0].get("name") or item.get("name")

        return {"requests": requests}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/student-leaves/{request_id}")
async def update_student_leave(
    request_id: str,
    status_update: LeaveStatusUpdate,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Approve or reject a student leave request."""
    try:
        existing = supabase.table("leave_requests").select("status").eq("id", request_id).execute()
        if not existing.data:
            raise HTTPException(status_code=404, detail="Request not found")

        current_status = existing.data[0].get("status", "pending_faculty")
        next_status = get_next_leave_status(current_status, "faculty", status_update.status == "approved")

        update_data = {
            "status": next_status,
            "updated_at": datetime.now().isoformat()
        }
        if status_update.rejection_reason:
            update_data["rejection_reason"] = status_update.rejection_reason

        result = supabase.table("leave_requests").update(update_data).eq("id", request_id).execute()
        if not result.data:
            raise HTTPException(status_code=404, detail="Request not found")

        return {"message": f"Leave request {next_status}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/my-leaves")
async def request_leave(
    leave: LeaveRequestCreate,
    current_user: CurrentUser = Depends(require_faculty)
):
    """Submit a leave request to HOD."""
    try:
        data = {
            "user_id": current_user.id,
            "role": "faculty",
            "leave_type": leave.leave_type,
            "from_date": leave.from_date.isoformat(),
            "to_date": leave.to_date.isoformat(),
            "reason": leave.reason,
            "status": "pending"
        }
        
        result = supabase.table("leave_requests").insert(data).execute()
        return {"message": "Leave request submitted successfully", "request": result.data[0]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/my-leaves")
async def get_my_leaves(current_user: CurrentUser = Depends(require_faculty)):
    """Get own leave history."""
    try:
        result = supabase.table("leave_requests").select("*").eq("user_id", current_user.id).order("created_at", desc=True).execute()
        faculty_result = supabase.table("faculty").select("name").eq("id", current_user.id).limit(1).execute()
        faculty_name = (faculty_result.data or [{}])[0].get("name") or current_user.email or "N/A"

        requests = result.data or []
        for item in requests:
            item["name"] = item.get("name") or faculty_name

        return {"requests": requests}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

